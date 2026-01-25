import openpyxl
from openpyxl.styles import Border, Side

def apply_excel_styling(file_path):
    """
    Applies thick borders to all cells containing data in the Excel file.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Define border style
        # "Thick" implies a visible, heavy border. 
        # 'medium' is often a good "thick" line in Excel that isn't the super-thick header style.
        # But user asked for "more thick", let's use 'medium' for inside and 'thick' for outside?
        # Or just 'medium' for everything is usually what people mean by "thick borders".
        # Let's try 'medium' for all for consistency as requested "outline and inline".
        
        border_style = Side(style='medium', color="000000")
        full_border = Border(left=border_style, 
                             right=border_style, 
                             top=border_style, 
                             bottom=border_style)
        
        # Iterate over all cells in the used range
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = full_border

        # Auto-adjust column widths (Optional but nice)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(file_path)
        return True
    except Exception as e:
        print(f"Error styling Excel file: {e}")
        return False
